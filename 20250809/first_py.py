from openpyxl import load_workbook

def calculate_averages():
    file_path = r"C:\HYJ\abcd_c\20250809\data.xlsx"
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        
        print("=" * 50)
        print("학생 성적 평균 계산")
        print("=" * 50)
        
        # 실제 헤더는 행1(인덱스1)의 열1부터 시작
        header = [rows[1][1], rows[1][2], rows[1][3], rows[1][4]]  # 이름, 국어, 영어, 수학
        print(f"컬럼: {header}")
        print("-" * 50)
        
        # 실제 학생 데이터는 행2부터 시작
        students = rows[2:]  # 홍길동, 이순신, 고기동
        
        averages = []
        
        for student in students:
            # 실제 데이터는 열1부터 시작 (열0은 None)
            name = student[1]      # 이름
            korean = student[2]    # 국어
            english = student[3]   # 영어
            math = student[4]      # 수학
            
            # 빈 행은 건너뛰기
            if not name or name is None:
                continue
                
            print(f"{name}: 국어({korean}), 영어({english}), 수학({math})")
            
            # 평균 계산
            avg = (korean + english + math) / 3
            averages.append(avg)
            
            print(f"  → 평균: {avg:.1f}점")
        
        print("-" * 50)
        if averages:
            total_avg = sum(averages) / len(averages)
            print(f"전체 평균: {total_avg:.1f}점")
            print(f"총 학생 수: {len(averages)}명")
            
            # 최고점, 최저점
            max_avg = max(averages)
            min_avg = min(averages)
            max_idx = averages.index(max_avg)
            min_idx = averages.index(min_avg)
            
            print(f"최고점: {students[max_idx][1]} ({max_avg:.1f}점)")
            print(f"최저점: {students[min_idx][1]} ({min_avg:.1f}점)")
        else:
            print("계산할 데이터가 없습니다.")
            
    except Exception as e:
        print(f"오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    calculate_averages()